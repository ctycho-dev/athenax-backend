"""
Validate Projects.xlsx against the Data Specs rules.

Usage:
    python scripts/validate_xlsx.py [path/to/file.xlsx] [sheet_name]

Exits 0 if clean, 1 if any errors are found.
"""

import re
import sys
from pathlib import Path

import openpyxl

XLSX_PATH = Path(__file__).parent.parent / "Projects.xlsx"

VALID_STAGES = {
    "Pre-Seed", "Seed", "Series A", "Series B",
    "Launched", "Beta", "Active", "Active Development", "Acquired / Operating",
}

FORBIDDEN_EMPTY = {"n/a", "-", "none"}

URL_COLUMNS = {"logo", "twitter", "website", "discord", "docs", "other link", "github"}

PLAIN_TEXT_COLUMNS = {"name", "short_desc", "description", "quality_badge", "category"}

# Minimum pipe-separated fields expected per semicolon entry
PACKED_MIN_FIELDS = {"team": 1, "voices": 4, "bounties": 4}

# Max length must match the DB column (VARCHAR) it's imported into
MAX_LENGTH = {
    "name": 150,
    "short_desc": 250,
    "quality_badge": 50,
    "email": 200,
    "logo": 500,
    "category": 100,
}

# Same as MAX_LENGTH but applied per ";"-separated entry (e.g. multiple subcategories)
MULTI_MAX_LENGTH = {"subcategory": 100}

EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
YEAR_RE = re.compile(r"^\d{4}$")
FUNDING_RE = re.compile(r"^\d+(\.\d+)?$")


def _check_row(row_num: int, name: str, row: dict) -> tuple[list[str], list[str]]:
    errors: list[str] = []
    warnings: list[str] = []
    prefix = f"row {row_num:<3} [{name}]"

    def err(col: str, msg: str) -> None:
        errors.append(f"  {prefix}  ERROR    {col}: {msg}")

    def warn(col: str, msg: str) -> None:
        warnings.append(f"  {prefix}  WARNING  {col}: {msg}")

    for col, val in row.items():
        if not col or not val:
            continue

        # Forbidden placeholder
        if val.strip().lower() in FORBIDDEN_EMPTY:
            err(col, f"leave blank instead of {val!r}")
            continue

        # Plain-text fields must not contain separators
        if col in PLAIN_TEXT_COLUMNS:
            if "|" in val:
                err(col, "must not contain '|'")
            if ";" in val:
                err(col, "must not contain ';'")

        # URL validation
        if col in URL_COLUMNS:
            if not val.startswith("https://"):
                err(col, f"must start with https://, got {val!r}")

        # Length limits (must match DB column size)
        if col in MAX_LENGTH and len(val) > MAX_LENGTH[col]:
            err(col, f"exceeds {MAX_LENGTH[col]} chars (got {len(val)})")

        if col in MULTI_MAX_LENGTH:
            limit = MULTI_MAX_LENGTH[col]
            for entry in (e.strip() for e in val.split(";")):
                if entry and len(entry) > limit:
                    err(col, f"entry {entry[:40]!r}… exceeds {limit} chars (got {len(entry)})")

        # Field-specific rules
        if col == "funding" and val:
            if not FUNDING_RE.match(val.strip()):
                err(col, f"digits only (no $, commas, spaces), got {val!r}")

        if col == "founded" and val:
            stripped = val.strip()
            if not YEAR_RE.match(stripped) or not (1900 <= int(stripped) <= 2099):
                err(col, f"must be a 4-digit year 1900–2099, got {val!r}")

        if col == "stage" and val:
            if val.strip() not in VALID_STAGES:
                err(col, f"{val!r} not valid — choose: {', '.join(sorted(VALID_STAGES))}")

        if col == "email" and val:
            if not EMAIL_RE.match(val.strip()):
                err(col, f"invalid email: {val!r}")

        # Packed columns (team / voices / bounties)
        if col in PACKED_MIN_FIELDS:
            min_fields = PACKED_MIN_FIELDS[col]
            entries = [e.strip() for e in val.split(";") if e.strip()]
            for i, entry in enumerate(entries, start=1):
                parts = [p.strip() for p in entry.split("|")]
                if col == "team" and not parts[0]:
                    warn(col, f"entry {i}: name is empty")
                if len(parts) < min_fields:
                    warn(col, f"entry {i}: expected {min_fields}+ pipe-separated fields, got {len(parts)}")

    return errors, warnings


def _cell(v) -> str:
    """Normalise an openpyxl cell value to a plain string.
    Whole-number floats (e.g. 2021.0) are returned as integers ('2021').
    """
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def validate(path: Path, sheet_name: str | None = None) -> int:
    if not path.exists():
        print(f"ERROR: file not found: {path}")
        return 1

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            wb.close()
            print(f"ERROR: sheet {sheet_name!r} not found — available: {', '.join(wb.sheetnames)}")
            return 1
        ws = wb[sheet_name]
    else:
        ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not all_rows:
        print("ERROR: file is empty")
        return 1

    headers = [str(h).strip().lower() if h is not None else "" for h in all_rows[0]]

    if "name" not in headers:
        print("ERROR: no 'name' column found in row 1")
        return 1

    all_errors: list[str] = []
    all_warnings: list[str] = []
    seen_slugs: dict[str, int] = {}
    data_rows = 0

    for row_num, raw in enumerate(all_rows[1:], start=2):
        row: dict[str, str] = {}
        for i, v in enumerate(raw):
            if i < len(headers) and headers[i]:
                row[headers[i]] = _cell(v)

        name = row.get("name", "").strip()
        if not name:
            continue  # skip blank rows silently

        data_rows += 1

        # Required field
        if not name:
            all_errors.append(f"  row {row_num:<3}           ERROR    name: must not be blank")
            continue

        errs, warns = _check_row(row_num, name, row)
        all_errors.extend(errs)
        all_warnings.extend(warns)

        # Duplicate slug check
        slug = re.sub(r"[^\w\s-]", "", name.lower().strip())
        slug = re.sub(r"[\s_-]+", "-", slug).strip("-")
        if slug in seen_slugs:
            all_warnings.append(
                f"  row {row_num:<3} [{name}]  WARNING  name: duplicate slug (same as row {seen_slugs[slug]})"
            )
        else:
            seen_slugs[slug] = row_num

    # Report
    sheet_label = f"{path.name} [{ws.title}]"
    print(f"\nValidated {data_rows} rows from {sheet_label}\n")

    if all_errors:
        print("--- ERRORS ---")
        for line in all_errors:
            print(line)
        print()

    if all_warnings:
        print("--- WARNINGS ---")
        for line in all_warnings:
            print(line)
        print()

    summary = f"{len(all_errors)} error(s), {len(all_warnings)} warning(s)"
    if all_errors:
        print(f"✗ {summary} — fix errors before uploading.")
        return 1

    print(f"✓ {summary} — file looks good.")
    return 0


if __name__ == "__main__":
    target = Path(sys.argv[1]) if len(sys.argv) > 1 else XLSX_PATH
    sheet = sys.argv[2] if len(sys.argv) > 2 else None
    sys.exit(validate(target, sheet))
